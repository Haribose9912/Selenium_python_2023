import openpyxl


# excel_file = "TestData/Logindata.xlsx"
# excel_worksheet = "Logintest"


# def test():
#     workbook = o.load_workbook(excel_file)
#     worksheet = workbook[excel_worksheet]
#
#     row_count = worksheet.max_row
#     col_count = worksheet.max_column
#     print("Rows: ",row_count,"Columns: ",col_count)
#     print("student=",worksheet.cell(2,2).value)
#     print("student=",worksheet.cell(3,2).value)
#     print("8333956091=",worksheet.cell(5,2).value)


def getRowCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColumnCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column


def readData(path, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=row_num, column=column_num).value


def writeData(path, sheetname, row_num, column_num, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(path)


#

def get_data_from_excel(path, sheetname):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook(sheetname)
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    for r in range(2, total_rows + 1):
        row_list = []
        for c in r(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
    return final_list
